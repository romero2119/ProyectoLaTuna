"""
UTILIDADES PARA EL SISTEMA DE INVENTARIO
- Exportación de reportes (PDF/Excel)
- Normalización de unidades de medida
- Cálculos de costos para recetas
"""
import io
from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.template.loader import render_to_string
import logging
from django.core.mail import send_mail
from django.conf import settings

CONVERSIONES_UNIDADES = {
    'kg': {'base': 'g', 'factor': 1000, 'tipo': 'masa'},      # 1 kg = 1000 g
    'g': {'base': 'g', 'factor': 1, 'tipo': 'masa'},
    'l': {'base': 'ml', 'factor': 1000, 'tipo': 'volumen'},    # 1 L = 1000 ml
    'ml': {'base': 'ml', 'factor': 1, 'tipo': 'volumen'},
    'm': {'base': 'cm', 'factor': 100, 'tipo': 'longitud'},    # 1 m = 100 cm
    'cm': {'base': 'cm', 'factor': 1, 'tipo': 'longitud'},
    'unidad': {'base': 'unidad', 'factor': 1, 'tipo': 'unidad'},
}

# Umbrales de stock por unidad de medida
UMBRALES_STOCK = {
    'kg': {'critico': 0.5, 'bajo': 2, 'normal': 5, 'texto': 'kilogramos'},
    'g': {'critico': 50, 'bajo': 200, 'normal': 500, 'texto': 'gramos'},
    'l': {'critico': 0.5, 'bajo': 2, 'normal': 5, 'texto': 'litros'},
    'ml': {'critico': 50, 'bajo': 200, 'normal': 500, 'texto': 'mililitros'},
    'm': {'critico': 1, 'bajo': 5, 'normal': 10, 'texto': 'metros'},
    'cm': {'critico': 10, 'bajo': 50, 'normal': 100, 'texto': 'centímetros'},
    'unidad': {'critico': 3, 'bajo': 10, 'normal': 20, 'texto': 'unidades'},
}

# ============================================
# 2. FUNCIONES DE NORMALIZACIÓN DE UNIDADES
# ============================================

def normalizar_cantidad(cantidad, unidad_origen):
    """
    Convierte una cantidad a su unidad base (gramos, mililitros, cm o unidades)
    
    Args:
        cantidad: Número a convertir
        unidad_origen: Unidad de origen (kg, g, l, ml, m, cm, unidad)
    
    Returns:
        Cantidad normalizada a la unidad base
    """
    if not cantidad or not unidad_origen:
        return 0
    
    if unidad_origen in CONVERSIONES_UNIDADES:
        return float(cantidad) * CONVERSIONES_UNIDADES[unidad_origen]['factor']
    return float(cantidad)

def obtener_precio_por_unidad_base(insumo_obj):
    """
    Calcula el precio por unidad base (ej: precio por gramo, por ml, por cm)
    
    Args:
        insumo_obj: Objeto insumo con campos stock, valor, unidad
    
    Returns:
        Precio por unidad base
    """
    if not insumo_obj or insumo_obj.stock == 0:
        return 0
    
    unidad = insumo_obj.unidad
    stock_normalizado = normalizar_cantidad(insumo_obj.stock, unidad)
    
    if stock_normalizado == 0:
        return 0
    
    # Precio por unidad base
    return float(insumo_obj.valor) / stock_normalizado

def calcular_costo_insumo(insumo_obj, cantidad_solicitada):
    """
    Calcula el costo REAL de un insumo para una receta
    
    IMPORTANTE: 
    - insumo_obj.valor = PRECIO UNITARIO (precio por kg, por g, por unidad, etc.)
    - insumo_obj.stock = CANTIDAD DISPONIBLE en esa unidad
    
    Ejemplo CORRECTO:
    Insumo: Queso
    Precio unitario: $8,000 por kg
    Stock: 6 kg disponibles
    Si la receta pide 30 kg:
    Costo = 30 kg × $8,000/kg = $240,000 ✅
    
    Args:
        insumo_obj: Objeto insumo
        cantidad_solicitada: Cantidad solicitada en la unidad del insumo
    
    Returns:
        Costo total redondeado a 2 decimales
    """
    if not insumo_obj or cantidad_solicitada <= 0:
        return 0
    
    # El valor ya es el precio por unidad
    precio_por_unidad = float(insumo_obj.valor)
    
    # Costo total = cantidad_solicitada × precio_por_unidad
    costo_total = float(cantidad_solicitada) * precio_por_unidad
    
    return round(costo_total, 2)

def formatear_cantidad_con_unidad(cantidad, unidad):
    """
    Formatea la cantidad mostrando la unidad apropiada
    Ej: 1000g -> 1 kg, 1500ml -> 1.5 L
    
    Args:
        cantidad: Número a formatear
        unidad: Unidad original
    
    Returns:
        String formateado con la unidad
    """
    if not cantidad:
        return f"0 {unidad}"
    
    if unidad == 'g' and cantidad >= 1000:
        return f"{cantidad/1000:.2f} kg"
    elif unidad == 'ml' and cantidad >= 1000:
        return f"{cantidad/1000:.2f} L"
    elif unidad == 'cm' and cantidad >= 100:
        return f"{cantidad/100:.2f} m"
    else:
        return f"{cantidad} {unidad}"

def obtener_estado_stock(insumo_obj):
    """
    Retorna el estado del stock según su unidad
    """
    if not insumo_obj:
        return {'clase': 'bg-secondary', 'texto': 'Sin datos', 'estado': 'unknown'}
    
    unidad = insumo_obj.unidad
    stock = insumo_obj.stock
    
    umbrales = UMBRALES_STOCK.get(unidad, {'critico': 3, 'bajo': 10, 'normal': 20})
    
    if stock == 0:
        return {
            'clase': 'bg-secondary',
            'texto': f'Sin stock',
            'estado': 'sin_stock',
            'stock_formateado': f"{stock} {unidad}"
        }
    elif stock < umbrales['critico']:
        return {
            'clase': 'bg-danger',
            'texto': f'{stock} {unidad} (Crítico)',
            'estado': 'critico',
            'stock_formateado': f"{stock} {unidad}"
        }
    elif stock <= umbrales['bajo']:
        return {
            'clase': 'bg-warning text-dark',
            'texto': f'{stock} {unidad} (Bajo)',
            'estado': 'bajo',
            'stock_formateado': f"{stock} {unidad}"
        }
    else:
        return {
            'clase': 'bg-success',
            'texto': f'{stock} {unidad} (Normal)',
            'estado': 'normal',
            'stock_formateado': f"{stock} {unidad}"
        }
def es_stock_bajo(insumo_obj):
    """
    Verifica si el stock está bajo según su unidad
    
    Args:
        insumo_obj: Objeto insumo
    
    Returns:
        Boolean indicando si el stock está bajo
    """
    if not insumo_obj:
        return False
    
    unidad = insumo_obj.unidad
    stock = insumo_obj.stock
    
    umbrales = UMBRALES_STOCK.get(unidad, {'bajo': 10})
    return stock <= umbrales['bajo']

# ============================================
# 3. FUNCIONES DE EXPORTACIÓN A PDF
# ============================================

def exportar_pdf(titulo, columnas, datos, nombre_archivo):
    """
    FUNCION PARA EXPORTAR DATOS A PDF USANDO WEASYPRINT
    
    Args:
        titulo: Titulo del reporte
        columnas: Lista de nombres de columnas
        datos: Lista de tuplas o diccionarios con los datos
        nombre_archivo: Nombre del archivo PDF a descargar
    
    Returns:
        HttpResponse con el PDF generado
    """
    
    # Crear contexto para el template
    contexto = {
        'titulo': titulo,
        'columnas': columnas,
        'datos': datos,
    }
    
    # Generar HTML desde el template
    html_string = render_to_string('reportes/reportes_pdf.html', contexto)
    
    # Crear documento PDF desde el HTML
    html_object = HTML(string=html_string, base_url='.')
    
    # Generar PDF en memoria
    pdf_bytes = html_object.write_pdf()
    
    # Crear respuesta HTTP con el PDF
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
    
    return response


def exportar_excel(titulo, columnas, datos, nombre_archivo):
    """
    FUNCIÓN PARA EXPORTAR DATOS A EXCEL USANDO OPENPYXL
    
    Args:
        titulo: Titulo del reporte
        columnas: Lista de nombres de columnas
        datos: Lista de tuplas o diccionarios con los datos
        nombre_archivo: Nombre del archivo Excel a descargar
    
    Returns:
        HttpResponse con el archivo Excel generado
    """
    
    # Crear un nuevo libro de Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte"
    
    # Configurar estilos para el título
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    # Agregar titulo
    worksheet.merge_cells('A1:' + chr(64 + len(columnas)) + '1')
    titulo_cell = worksheet['A1']
    titulo_cell.value = titulo
    titulo_cell.font = title_font
    titulo_cell.fill = title_fill
    titulo_cell.alignment = title_alignment
    worksheet.row_dimensions[1].height = 25

    # Configurar estilos para los encabezados
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Agregar encabezados de columnas
    for col_num, columna in enumerate(columnas, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = columna
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    worksheet.row_dimensions[3].height = 20
    
    # Configurar estilos para los datos
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Agregar datos al Excel
    data_fill_alternated = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    for row_num, fila in enumerate(datos, 4):
        # Convertir diccionario a tupla si es necesario
        if isinstance(fila, dict):
            valores = [fila.get(col.lower().replace(' ', '_'), '') for col in columnas]
        else:
            valores = fila
        
        # Llenar las celdas con datos
        for col_num, valor in enumerate(valores, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = valor
            cell.alignment = data_alignment
            cell.border = data_border
            
            # Colorear filas alternas para mejor legibilidad
            if (row_num - 4) % 2 == 0:
                cell.fill = data_fill_alternated
    
    # Ajustar ancho de columnas automaticamente
    for col_num, columna in enumerate(columnas, 1):
        max_length = len(str(columna))
        column_letter = chr(64 + col_num)
        
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
        # Para columnas más allá de la Z (si hay más de 26 columnas)
        if col_num > 26:
            from openpyxl.utils import get_column_letter
            column_letter = get_column_letter(col_num)
        
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Crear respuesta HTTP con el Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    
    # Guardar el libro en la respuesta
    workbook.save(response)
    
    return response


# ============================================
# 5. FUNCIONES DE EXPORTACIÓN ESPECIALIZADAS
# ============================================

def exportar_insumos_con_estado(insumos_queryset, formato='excel'):
    """
    Exporta insumos con su estado de stock calculado
    """
    columnas = ['ID', 'Nombre', 'Categoría', 'Stock', 'Unidad', 'Estado', 'Precio Unitario', 'Valor Total Stock']
    
    datos = []
    for insumo in insumos_queryset:
        estado = obtener_estado_stock(insumo)
        valor_total_stock = insumo.stock * insumo.valor  # Precio unitario × cantidad
        
        datos.append({
            'id': insumo.id_insumo,
            'nombre': insumo.nombre,
            'categoría': insumo.categoria.nombre if insumo.categoria else 'Sin categoría',
            'stock': f"{insumo.stock} {insumo.unidad}",
            'unidad': insumo.unidad,
            'estado': estado['texto'],
            'precio_unitario': f"${insumo.valor:,.2f}",
            'valor_total_stock': f"${valor_total_stock:,.2f}"
        })
    
    titulo = "Reporte de Inventario - Insumos"
    nombre_archivo = f"insumos_{formato}"
    
    if formato == 'excel':
        return exportar_excel(titulo, columnas, datos, nombre_archivo)
    else:
        return exportar_pdf(titulo, columnas, datos, nombre_archivo)

def exportar_recetas_con_costos(recetas_queryset, formato='excel'):
    """
    Exporta recetas con sus costos calculados
    
    Args:
        recetas_queryset: QuerySet de recetas
        formato: 'excel' o 'pdf'
    
    Returns:
        HttpResponse con el archivo generado
    """
    columnas = ['ID', 'Plato', 'Cantidad de Insumos', 'Costo Total', 'Precio Sugerido (con IVA)']
    
    datos = []
    for receta in recetas_queryset:
        costo_total = 0
        cantidad_insumos = receta.detalles.count()
        
        for detalle in receta.detalles.all():
            costo_insumo = calcular_costo_insumo(detalle.insumo, detalle.cantidad)
            costo_total += costo_insumo
        
        iva = costo_total * 0.19
        precio_sugerido = costo_total + iva
        
        datos.append({
            'id': receta.id,
            'plato': receta.plato.nombre,
            'cantidad_insumos': cantidad_insumos,
            'costo_total': f"${costo_total:,.2f}",
            'precio_sugerido': f"${precio_sugerido:,.2f}"
        })
    
    titulo = "Reporte de Recetas - Análisis de Costos"
    nombre_archivo = f"recetas_{formato}"
    
    if formato == 'excel':
        return exportar_excel(titulo, columnas, datos, nombre_archivo)
    else:
        return exportar_pdf(titulo, columnas, datos, nombre_archivo)
    
